import random
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from calendar import monthrange
import os
import psycopg2
from psycopg2.extras import RealDictCursor

class XlsxGenerator:
    def __init__(self, db_config):
        self.conn = psycopg2.connect(**db_config)

        self.COST_ITEMS = {
            "Топливо": ["Лукойл", "Газпромнефть", "Татнефть", "Shell", "Роснефть"],
            "Ремонт (запчасти)": ["Автомир", "Exist.ru", "Авто-Люкс", "Emex"],
            "Ремонт (работа)": ["Автосервис №1", "СТО Профи", "Автотехцентр"],
            "Платные дороги": ["Автодор", "М-11 Нева"],
            "Мойка": ["Чистый борт", "Мойка 24/7", "Аква-джет"],
            "Страховка": ["АльфаСтрах", "СберСтрах", "Росгосстрах", "Ингосстрах"],
            "Шиномонтаж": ["Колесо.ру", "Шиномонтаж-Юг", "АвтоШина"],
            "Заправка кондиционера": ["Климат-Сервис"],
            "Масло/техжидкости": ["Масло-Маркет", "Лукойл"],
            }

        self.EXTRA_CONTRACTORS = {
            "Топливо": ["АЗС №5", "АЗС №12", "ТРК №3", "АЗС на Лесной"],
            "Ремонт (запчасти)": ["Деталь-Маркет", "Автозапчасти32"],
            "Платные дороги": ["М-4 Дон", "М-5 Урал"],
            "Прочее": ["Автомагазин", "—"],
            }

        self.NOTES = {
            "Топливо": ["Заправка АИ-95", "Заправка дизелем", "Полный бак", "Обычная заправка"],
            "Ремонт (запчасти)": ["Замена ремня ГРМ", "Тормозные колодки", "Фильтр масляный", "Свечи зажигания", "Ремкомплект"],
            "Ремонт (работа)": ["Диагностика", "Замена масла", "Регулировка развал-схождения", "Замена ремня"],
            "Платные дороги": ["Проезд по трассе", "Оплата участка", "Транспондер"],
            "Мойка": ["Мойка кузова", "Химчистка салона", "Мойка двигателя", "Комплексная мойка"],
            "Страховка": ["ОСАГО", "КАСКО"],
            "Шиномонтаж": ["Сезонная смена шин", "Ремонт прокола", "Балансировка колес"],
            "Заправка кондиционера": ["Заправка фреоном", "Диагностика кондиционера"],
            "Масло/техжидкости": ["Замена масла", "Тосол", "Тормозная жидкость"],
            "Прочее": ["Щетки стеклоочистителя", "Ароматизатор", "Антифриз", "Омывайка"],
            }

    #Получение водителей из Базы постгрес для генерации свящанных данных
    def get_vehicles_with_drivers_from_db(self):
        try:
            cursor = self.conn.cursor(cursor_factory=RealDictCursor)
            
            query = """
                SELECT DISTINCT
                    v.plate_number,
                    v.model,
                    d.full_name as driver_name
                FROM vehicles v
                JOIN sale_headers sh ON sh.vehicle_id = v.id
                JOIN drivers d ON sh.driver_id = d.id
                WHERE v.is_active = true
                AND d.is_active = true
                ORDER BY v.plate_number
            """
            
            cursor.execute(query)
            results = cursor.fetchall()
            
            cursor.close()
            self.conn.close()
            
            vehicles_with_drivers = []
            for row in results:
                vehicles_with_drivers.append({
                    'plate_number': row['plate_number'],
                    'driver_name': row['driver_name']
                })
            
            return vehicles_with_drivers
        except Exception as e:
            print(f"Ошибка при получении данных из БД: {e}")
            return []

    #Генерация случайной даты в месяце
    def generate_random_date_in_month(self, year, month):
        last_day = monthrange(year, month)[1]
        day = random.randint(1, last_day)
        return f"{day:02d}.{month:02d}"

    #Генерация одной записи
    def generate_random_record(self, date, vehicle_info, driver_name, cost_item, contractors):
        if cost_item in self.EXTRA_CONTRACTORS:
            contractor_pool = contractors + self.EXTRA_CONTRACTORS[cost_item]
        else:
            contractor_pool = contractors
        
        contractor = random.choice(contractor_pool) if contractor_pool else "-"
        
        doc_types = ["Чек", "Накл.", "Акт", "Счет", "Полис", "Квитанция"]
        doc_num = f"{random.choice(doc_types)} {random.randint(1, 999)}"
        
        if cost_item == "Топливо":
            amount = random.randint(1500, 7000)
        elif cost_item == "Страховка":
            amount = random.choice([5000, 8000, 12000, 15000, 20000])
        elif cost_item in ["Ремонт (запчасти)", "Ремонт (работа)"]:
            amount = random.randint(3000, 25000)
        elif cost_item == "Мойка":
            amount = random.randint(500, 2500)
        elif cost_item == "Платные дороги":
            amount = random.randint(200, 1800)
        else:
            amount = random.randint(300, 10000)
        
        if cost_item in self.NOTES:
            note = random.choice(self.NOTES[cost_item])
        else:
            note = "-"
        
        return [date, vehicle_info, driver_name, cost_item, contractor, doc_num, amount, note]
    
    #Заполнение таблицы записями за месяц
    def generate_report_for_month(self, year, month, vehicles_with_drivers, records_per_vehicle=6):
        all_records = []
        
        for vehicle in vehicles_with_drivers:
            plate_number = vehicle['plate_number']
            driver_name = vehicle['driver_name']
            
            vehicle_info = f"{plate_number}"
            
            num_records = random.randint(records_per_vehicle - 2, records_per_vehicle + 2)
            
            for i in range(num_records):
                record_date = self.generate_random_date_in_month(year, month)
                
                cost_items = list(self.COST_ITEMS.keys())
                weights = [0.6, 0.05, 0.03, 0.1, 0.1, 0.03, 0.02, 0.02, 0.03, 0.02]
                while len(weights) < len(cost_items):
                    weights.append(0.01)
                weights = weights[:len(cost_items)]
                total = sum(weights)
                weights = [w/total for w in weights]
                
                cost_item = random.choices(cost_items, weights=weights)[0]
                contractors = self.COST_ITEMS[cost_item]
                
                record = self.generate_random_record(
                    record_date, vehicle_info, driver_name, cost_item, contractors
                )
                all_records.append(record)
        
        all_records.sort(key=lambda x: int(x[0].split(".")[0]))
        
        return all_records

    #Сохранение в эксель
    def save_to_excel(self, records, year, month, output_dir="data/raw/excel/"):
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        month_names = {
            1: "январь", 2: "февраль", 3: "март", 4: "апрель",
            5: "май", 6: "июнь", 7: "июль", 8: "август",
            9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь"
        }
        filename = f"{month_names[month]}_{year}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = f"{month_names[month].capitalize()} {year}"
        
        headers = ["Дата", "Гос. номер", "Водитель", "Статья затрат", 
                "Контрагент (кому заплатили)", "Номер документа", 
                "Сумма, руб.", "Примечание"]
        sheet.append(headers)
        
        for record in records:
            sheet.append(record)
        
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        column_widths = [10, 12, 12, 18, 28, 15, 14, 35]
        for i, width in enumerate(column_widths, 1):
            sheet.column_dimensions[chr(64 + i)].width = width
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in sheet.iter_rows(min_row=1, max_row=len(records)+1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                if cell.column == 7:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00" ₽"'
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        
        total_row = len(records) + 2
        total_sum = sum(record[6] for record in records)
        
        sheet.cell(row=total_row, column=6, value="ИТОГО:").font = Font(bold=True, size=11)
        sheet.cell(row=total_row, column=7, value=total_sum).number_format = '#,##0.00" ₽"'
        sheet.cell(row=total_row, column=6).alignment = Alignment(horizontal="right")
        
        sheet.freeze_panes = "A2"
        
        wb.save(filepath)
        print(f"Сгенерирован: {filepath}")
        print(f"Записей: {len(records)}, Сумма: {total_sum:,.2f} руб.")
        
        return total_sum

    #Главная функция
    def generate_report_for_period(self, start_date, end_date, records_per_vehicle=6, output_dir="data/raw/excel"):
        print(f"\nГЕНЕРАЦИЯ ОТЧЕТОВ ЗА ПЕРИОД\n")
        print(f"Период: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}")
        print(f"Записей на машину в месяц: ~{records_per_vehicle}")
        print(f"Папка сохранения: {output_dir}/\n")
        
        vehicles_with_drivers = self.get_vehicles_with_drivers_from_db()
        
        if not vehicles_with_drivers:
            print("Ошибка: Не удалось получить данные из БД")
            return
        
        print(f"Загружено {len(vehicles_with_drivers)} автомобилей с водителями из БД\n")
        
        current_date = start_date.replace(day=1)
        end_date_first = end_date.replace(day=1)
        
        generated_files = []
        
        while current_date <= end_date_first:
            year = current_date.year
            month = current_date.month
            
            print(f"Генерация за {month:02d}.{year}...")
            
            records = self.generate_report_for_month(year, month, vehicles_with_drivers, records_per_vehicle)
            
            if records:
                self.save_to_excel(records, year, month, output_dir)
                generated_files.append(f"{month:02d}.{year}")
            
            if month == 12:
                current_date = current_date.replace(year=year + 1, month=1)
            else:
                current_date = current_date.replace(month=month + 1)
            
            print()
        
        print(f"ГОТОВО! Сгенерировано {len(generated_files)} файлов")
        print(f"Файлы сохранены в папке: {output_dir}/")